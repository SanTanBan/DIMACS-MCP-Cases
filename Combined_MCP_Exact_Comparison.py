from gurobipy import *
import winsound
import openpyxl

# "SanBan.b",
# "brock200_2.b",
# "brock200_4.b",
# "brock400_2.b",
# "brock400_4.b",
# "brock800_2.b",
# "brock800_4.b",
# "C125.9.clq",
# "C250.9.clq",
# "C500.9.clq",
# "C1000.9.clq",

filenames=[
"C2000.5.clq",
"C4000.5.clq",
"DSJC500.5.clq",
"DSJC1000.5.clq",
"gen200_p0.9_44.b",
"gen200_p0.9_55.b",
"gen400_p0.9_55.b",
"gen400_p0.9_65.b",
"gen400_p0.9_75.b",
"hamming10-4.clq",
"hamming8-4.clq",
"keller4.clq",
"keller5.clq",
"MANN_a27.clq.b",
"MANN_a45.clq.b",
"p_hat1500-1.clq",
"p_hat1500-2.clq",
"p_hat1500-3.clq",
"p_hat300_1.clq",
"p_hat300-2.clq",
"p_hat300-3.clq",
"p_hat700-1.clq",
"p_hat700-2.clq",
"p_hat700-3.clq"]

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.create_sheet("Result_3333_sec")

# Write headers
ws.cell(row=1, column=1, value="FileName")
ws.cell(row=1, column=2, value="Min_Node_Number")
ws.cell(row=1, column=3, value="Max_Node_Number")
ws.cell(row=1, column=4, value="Number_of_Edges")

ws.cell(row=1, column=6, value="Objective1: Max Clique Value")
ws.cell(row=1, column=7, value="Time for Book_Formulation in seconds")
ws.cell(row=1, column=8, value="OptimalityGap")
ws.cell(row=1, column=9, value="Best_Bound")

ws.cell(row=1, column=11, value="Objective2: Max Clique Value")
ws.cell(row=1, column=12, value="Time for My_New_Formulation in seconds")
ws.cell(row=1, column=13, value="OptimalityGap")
ws.cell(row=1, column=14, value="Best_Bound")

ws.cell(row=1, column=16, value="Objective3: Max Clique Value")
ws.cell(row=1, column=17, value="Time for Combined_Formulation in seconds")
ws.cell(row=1, column=18, value="OptimalityGap")
ws.cell(row=1, column=19, value="Best_Bound")

# Smallest_clq="C125.9.clq"
# Smallest_b="brock200_2.b"
#Smallest_b=

row_count = 2
for single_file_name in filenames:
    row_count += 1
        
    # Using readlines()
    file_full = open(single_file_name, 'r')
    Lines = file_full.readlines()
    edges=[]
    min_Node_value=999999999
    max_Node_value=0


    # Strips the newline character
    for line in Lines:
        #print("Line{}: {}".format(count, line.strip()))
        #print(type(line))

        line_words=line.split()
        if line_words[0]=='e':
            Node1=int(line_words[1])
            Node2=int(line_words[2])
            
            edges.append((Node1,Node2))
            
            if Node1 > max_Node_value:
                max_Node_value=Node1
            if Node2 > max_Node_value:
                max_Node_value=Node2

            if Node1 < min_Node_value:
                min_Node_value=Node1
            if Node2 < min_Node_value:
                min_Node_value=Node2

    ws.cell(row=row_count, column=1, value=single_file_name)
    ws.cell(row=row_count, column=2, value=min_Node_value)
    ws.cell(row=row_count, column=3, value=max_Node_value)
    ws.cell(row=row_count, column=4, value=len(edges))

    print("Nodes available: ",min_Node_value," -to- ",max_Node_value)
    node_names=range(min_Node_value,max_Node_value+1)

    # for i in node_names:
    #     #print("\n")
    #     drt=''
    #     for j in node_names:
    #         if j!=i:
    #             if ((i,j) in edges or (j,i) in edges):
    #                 drt+="1"
    #             else:
    #                 drt+="0"
    #         else:
    #             drt+="0"
    #     print(drt)


    #The Maximum Clique Problem (MCP) is a well-known NP-hard problem in computer science that involves finding the largest complete subgraph in a given undirected graph. Due to its computational complexity, it is difficult to solve large-scale instances of this problem exactly.

    #The largest Maximum Clique Problem ever solved to optimality, according to the literature, involves a graph with 1,192,780 vertices and 46,994,362 edges. The solution was found by a team of researchers from the University of Colorado and the University of Edinburgh in 2016 using a parallel branch-and-bound algorithm called "Mace" that was specifically designed for solving the MCP. The computation took approximately 1,100 CPU years, which was achieved using a high-performance computing cluster with 1,000 nodes, each with 28 Intel Xeon cores.

    #It is worth noting that while the above result represents the largest MCP solved to optimality reported in the literature, it is possible that larger instances have been solved using heuristic or approximation algorithms. Additionally, the time required to solve an MCP can vary widely depending on the specific algorithm and hardware used, as well as the size and structure of the graph being analyzed.



    print("**********************************************")
    print("Book_Formulation for file: ",single_file_name)
    print("**********************************************")
    # BOOK FORMULATION
    y={}
    # Set the problem
    mdl=Model("MCP_Book_Formulation")
    mdl.setParam('TimeLimit', 3333) # 3600 seconds is 1 hour

    for i in node_names:
        y[i] = mdl.addVar(vtype=GRB.BINARY, name="y%s" % (i))
        mdl.update()

    # Equation 1
    mdl.setObjective(quicksum(y[i] for i in node_names),GRB.MAXIMIZE)

    # Equation 2
    mdl.addConstrs(y[j] + y[i] <=1 for i in node_names for j in node_names if j>i if (j,i) not in edges)
    mdl.update()

    # Solve the Problem using default Gurobi settings
    mdl.optimize()
    #winsound.Beep(500, 1000) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds

    print("**********************************************")
    OptimalityGap=mdl.MIPGap
    print("Final MIP Gap value: %f" % OptimalityGap)
    best_bound=mdl.ObjBound
    run_time=mdl.Runtime
    print("Best Bound found: ",best_bound," ; Found in ",run_time," seconds")
    Solutions_Found=mdl.SolCount

    if Solutions_Found:
        objec_val=mdl.getObjective().getValue() #mdl.objVal
        print("Objective Value: ",objec_val)
        
        ws.cell(row=row_count, column=6, value=objec_val)
        ws.cell(row=row_count, column=7, value=run_time)
        ws.cell(row=row_count, column=8, value=OptimalityGap)
        ws.cell(row=row_count, column=9, value=best_bound)

    # Save workbook to file
    wb.save("Foumulation_Results.xlsx")
    mdl.dispose()
    print("**********************************************")




    # for i in node_names:
    #     #print("\n")
    #     drt=''
    #     for j in node_names:
    #         if j!=i:
    #             if ((i,j) in edges or (j,i) in edges):
    #                 drt+="1"
    #             else:
    #                 drt+="0"
    #         else:
    #             drt+="0"
    #     print(drt)


    print("**********************************************")
    print("My_New_Formulation for file: ",single_file_name)
    print("**********************************************")
    # MY NEW FORMULATION
    M=999999999 # Too Large M value is giving wrong results
    y={}
    # Set the problem
    mdl=Model("MCP_My_New_Formulation")
    mdl.setParam('TimeLimit', 3333) # 3600 seconds is 1 hour

    for i in node_names:
        y[i] = mdl.addVar(vtype=GRB.BINARY, name="y%s" % (i))
        mdl.update()

    x = mdl.addVar(vtype=GRB.CONTINUOUS, name="x")
    mdl.update()

    # Equation 1
    mdl.setObjective(x,GRB.MAXIMIZE)

    # Equation 2
    mdl.addConstr(x == quicksum(y[i] for i in node_names))
    mdl.update()

    # Equation 3
    mdl.addConstrs(quicksum(y[j] for j in node_names if j!=i if ((i,j) in edges or (j,i) in edges)) >= x-M*(1-y[i])-1 for i in node_names)
    mdl.update()

    # Equation 4
    mdl.addConstrs(quicksum(y[j] for j in node_names if j!=i if ((i,j) in edges or (j,i) in edges)) <= x+M*(1-y[i])-1 for i in node_names)
    mdl.update()

    # Solve the Problem using default Gurobi settings
    mdl.optimize()
    #winsound.Beep(500, 1000) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds

    print("**********************************************")
    OptimalityGap=mdl.MIPGap
    print("Final MIP Gap value: %f" % OptimalityGap)
    best_bound=mdl.ObjBound
    run_time=mdl.Runtime
    print("Best Bound found: ",best_bound," ; Found in ",run_time," seconds")

    Solutions_Found=mdl.SolCount

    if Solutions_Found:
        objec_val=mdl.getObjective().getValue() #mdl.objVal
        print("Objective Value: ",objec_val)

        ws.cell(row=row_count, column=11, value=objec_val)
        ws.cell(row=row_count, column=12, value=run_time)
        ws.cell(row=row_count, column=13, value=OptimalityGap)
        ws.cell(row=row_count, column=14, value=best_bound)        
    
    # Save workbook to file
    wb.save("Foumulation_Results.xlsx")
    mdl.dispose()
    print("**********************************************")





    print("**********************************************")
    print("Combined_Formulation for file: ",single_file_name)
    print("**********************************************")
    # MY NEW FORMULATION COMBINED WITH BOOK FORMULATION
    M=999999999 # Too Large M value is giving wrong results
    y={}
    # Set the problem
    mdl=Model("MCP_Combined_Formulation")
    mdl.setParam('TimeLimit', 3333) # 3600 seconds is 1 hour

    for i in node_names:
        y[i] = mdl.addVar(vtype=GRB.BINARY, name="y%s" % (i))
        mdl.update()

    # Equation 1
    mdl.setObjective(quicksum(y[i] for i in node_names),GRB.MAXIMIZE)

    # Equation 2
    mdl.addConstrs(y[j] + y[i] <=1 for i in node_names for j in node_names if j>i if (j,i) not in edges)
    mdl.update()

    # Equation 3
    mdl.addConstrs(quicksum(y[j] for j in node_names if j!=i if ((i,j) in edges or (j,i) in edges)) >= quicksum(y[k] for k in node_names if k!=i) - M*(1-y[i]) for i in node_names)
    mdl.update()

    # Equation 4
    mdl.addConstrs(quicksum(y[j] for j in node_names if j!=i if ((i,j) in edges or (j,i) in edges)) <= quicksum(y[k] for k in node_names if k!=i) + M*(1-y[i]) for i in node_names)
    mdl.update()

    # Solve the Problem using default Gurobi settings
    mdl.optimize()
    #winsound.Beep(500, 1000) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds

    print("**********************************************")
    OptimalityGap=mdl.MIPGap
    print("Final MIP Gap value: %f" % OptimalityGap)
    best_bound=mdl.ObjBound
    run_time=mdl.Runtime
    print("Best Bound found: ",best_bound," ; Found in ",run_time," seconds")

    Solutions_Found=mdl.SolCount

    if Solutions_Found:
        objec_val=mdl.getObjective().getValue() #mdl.objVal
        print("Objective Value: ",objec_val)

        ws.cell(row=row_count, column=16, value=objec_val)
        ws.cell(row=row_count, column=17, value=run_time)
        ws.cell(row=row_count, column=18, value=OptimalityGap)
        ws.cell(row=row_count, column=19, value=best_bound)        
    
    # Save workbook to file
    wb.save("Foumulation_Results.xlsx")
    mdl.dispose()
    print("**********************************************")


del wb["Sheet"]