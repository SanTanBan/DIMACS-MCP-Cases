from gurobipy import *
import winsound
import openpyxl

filenames=["brock200_2.b",
"brock200_4.b",
"brock400_2.b",
"brock400_4.b",
"brock800_2.b",
"brock800_4.b",
"C125.9.clq",
"C250.9.clq",
"C500.9.clq",
"C1000.9.clq",
"C2000.5.clq",
"C4000.5.clq",
"DSJC500.5.clq",
"DSJC1000.5.clq",
"gen200_p0.9_44.b",
"gen200_p0.9_55.b",
"gen400_p0.9_55.b",
"gen400_p0.9_65.b",
"gen400_p0.9_75.b",
"hamming10-4.clq",
"hamming8-4.clq",
"keller4.clq",
"keller5.clq",
"MANN_a27.clq.b",
"MANN_a45.clq.b",
"p_hat1500-1.clq",
"p_hat1500-2.clq",
"p_hat1500-3.clq",
"p_hat300_1.clq",
"p_hat300-2.clq",
"p_hat300-3.clq",
"p_hat700-1.clq",
"p_hat700-2.clq",
"p_hat700-3.clq"]


filenames=["brock200_2.b",
"p_hat700-3.clq"]

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.create_sheet("Results")

# Write headers
ws.cell(row=1, column=1, value="FileName")
ws.cell(row=1, column=2, value="Objective: Max Clique Value")
ws.cell(row=1, column=3, value="Time for Book_Formulation in seconds")


# Smallest_clq="C125.9.clq"
# Smallest_b="brock200_2.b"
#Smallest_b="SanBan.b"

results={}
row_count = 1
for single_file_name in filenames:
    row_count += 1

    # Using readlines()
    file1 = open(single_file_name, 'r')
    Lines = file1.readlines()
    edges=[]
    min_Node_value=9999999999
    max_Node_value=0


    # Strips the newline character
    for line in Lines:
        
        #print("Line{}: {}".format(count, line.strip()))
        #print(type(line))

        line_words=line.split()
        if line_words[0]=='e':
            Node1=int(line_words[1])
            Node2=int(line_words[2])
            
            edges.append((Node1,Node2))
            
            if Node1 > max_Node_value:
                max_Node_value=Node1
            if Node2 > max_Node_value:
                max_Node_value=Node2

            if Node1 < min_Node_value:
                min_Node_value=Node1
            if Node2 < min_Node_value:
                min_Node_value=Node2

        # if count >= 55:
        #     break

    #print(edges)

    # print("hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh")
    # if (3,9) in edges or (9,3) in edges:
    #     print(1)
    # else:
    #     print(0)

    print("Nodes available: ",min_Node_value," -to- ",max_Node_value)
    node_names=range(min_Node_value,max_Node_value+1)

    # for i in node_names:
    #     print(i)


    # for i in node_names:
    #     #print("\n")
    #     drt=''
    #     for j in node_names:
    #         if j!=i:
    #             if ((i,j) in edges or (j,i) in edges):
    #                 drt+="1"
    #             else:
    #                 drt+="0"
    #         else:
    #             drt+="0"
    #     print(drt)


    #"""

    M=9999999999
    y={}
    # Set the problem
    mdl=Model("MCP_Book_Formulation")
    #mdl.setParam('TimeLimit', 999) # 3600 seconds is 1 hour

    for i in node_names:
        y[i] = mdl.addVar(vtype=GRB.BINARY, name="y%s" % (i))
        mdl.update()

    # Equation 1
    mdl.setObjective(quicksum(y[i] for i in node_names),GRB.MAXIMIZE)


    # Equation 2
    mdl.addConstrs(y[j] + y[i] <=1 for i in node_names for j in node_names if j>i if (j,i) not in edges)
    mdl.update()


    #The Maximum Clique Problem (MCP) is a well-known NP-hard problem in computer science that involves finding the largest complete subgraph in a given undirected graph. Due to its computational complexity, it is difficult to solve large-scale instances of this problem exactly.

    #The largest Maximum Clique Problem ever solved to optimality, according to the literature, involves a graph with 1,192,780 vertices and 46,994,362 edges. The solution was found by a team of researchers from the University of Colorado and the University of Edinburgh in 2016 using a parallel branch-and-bound algorithm called "Mace" that was specifically designed for solving the MCP. The computation took approximately 1,100 CPU years, which was achieved using a high-performance computing cluster with 1,000 nodes, each with 28 Intel Xeon cores.

    #It is worth noting that while the above result represents the largest MCP solved to optimality reported in the literature, it is possible that larger instances have been solved using heuristic or approximation algorithms. Additionally, the time required to solve an MCP can vary widely depending on the specific algorithm and hardware used, as well as the size and structure of the graph being analyzed.


    # Solve the Problem using default Gurobi settings
    mdl.optimize()
    #winsound.Beep(500, 1000) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds

    OptimalityGap=mdl.MIPGap
    print("Final MIP Gap value: %f" % OptimalityGap)
    best_bound=mdl.ObjBound
    run_time=mdl.Runtime
    print("Best Bound found: ",best_bound," ; Found in ",run_time," seconds")

    Solutions_Found=mdl.SolCount

    if Solutions_Found:
        objec_val=mdl.getObjective().getValue() #mdl.objVal
        print("Objective Value: ",objec_val)
        results[single_file_name]=(objec_val,run_time)

        ws.cell(row=row_count, column=1, value=single_file_name)
        ws.cell(row=row_count, column=2, value=objec_val)
        ws.cell(row=row_count, column=3, value=run_time)



    # Save workbook to file
    wb.save("Book_Foumulation.xlsx")

del wb["Sheet"]
